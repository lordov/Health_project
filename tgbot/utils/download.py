import pandas as pd
import tempfile
import os


from aiogram.types import FSInputFile, CallbackQuery
from aiogram_dialog import DialogManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from tgbot.database.orm_query import orm_get_report
from tgbot.utils.logger_config import logging


async def download_xlsx(callback: CallbackQuery, dialog_manager: DialogManager):
    session = dialog_manager.middleware_data.get('session')
    # Инициализация логгера для этого модуля
    logger = logging.getLogger(__name__)

    # Логирование начала процесса скачивания xlsx
    logger.info("Начало процесса загрузки xlsx файла.")

    try:
        # Получение данных из базы данных
        report_data = await orm_get_report(session)

        # Преобразование данных в DataFrame с помощью pandas
        df = pd.DataFrame(report_data, columns=[
            'created', 'user_id', 'full_name', 'phone', 'operation_date',
            'answer1', 'answer2', 'answer3', 'answer4', 'answer5',
            'answer6', 'answer7', 'answer8', 'answer9', 'answer10',
            'answer11', 'answer12', 'answer13', 'answer14',
            'imt', 'self_assessment', 'pain_scale', 'number_of_survay'
        ])

        # Изменение названий текущих столбцов
        new_columns = {
            'created': 'Дата прохождения опроса',
            'user_id': 'ID пользователя',
            'full_name': 'ФИО',
            'phone': 'Номер телефона',
            'operation_date': 'Дата операции',
            'answer1': 'Боль в области оперативного вмешательства на грудной клетке',
            'answer2': 'Боль в области оперативного вмешательства на нижней конечности (ноге) после  проведённого АКШ',
            'answer3': 'Покраснение кожи в области послеоперационного рубца',
            'answer4': 'Подвижность в области грудины (щелчки)',
            'answer5': 'Выделение из послеоперационной раны на грудной клетке',
            'answer6': 'Повышение теммпературы тела (лихорадка)',
            'answer7': 'Сахарный диабет',
            'answer8': 'Гликемия утром',
            'answer9': 'Гликемия вечером',
            'answer10': 'ХОБЛ',
            'answer11': 'Курение',
            'answer12': 'Количество выкуриваемых пачек в день',
            'answer13': 'Кашель',
            'answer14': 'Одышка',
            'imt': 'ИМТ',
            'self_assessment': 'Субъективная оценка состояния здоровья',
            'pain_scale': 'ВАШ',
            'number_of_survay': '№ Опроса'
        }
        df.rename(columns=new_columns, inplace=True)

        # Транспонирование DataFrame
        df_transposed = df.transpose()

        # Создание временного xlsx-файла на диске
        temp_xlsx_fd, temp_xlsx_path = tempfile.mkstemp(suffix='.xlsx')
        with os.fdopen(temp_xlsx_fd, 'wb') as temp_xlsx_file:
            # Запись данных в xlsx-файл
            df_transposed.to_excel(temp_xlsx_file, index=True, header=False)

        # Открытие созданного файла для работы с стилями
        wb = load_workbook(temp_xlsx_path)
        ws = wb.active

        # Применение стиля выравнивания по левому краю для всех ячеек
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='left')

        # Сохранение изменений в файл
        wb.save(temp_xlsx_path)

        # Создание FSInputFile для отправки файла в телеграм
        xlsx_file = FSInputFile(temp_xlsx_path, filename='report.xlsx')

        # Отправка xlsx-файла пользователю
        await callback.message.answer_document(xlsx_file)

        # Логирование успешной отправки xlsx
        logger.info("Отправка xlsx файла завершена успешно.")

    except Exception as e:
        # Логирование ошибки, если что-то пошло не так
        logger.exception(f"Произошла ошибка при скачивании xlsx файла: {e}")

    finally:
        # Удаление временного файла
        os.unlink(temp_xlsx_path)
